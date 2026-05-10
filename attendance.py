from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

FILE_NAME = "attendance.xlsx"

def mark_attendance(name):
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Attendance"

        
        sheet.append(["Name", "Date", "Time"])

        wb.save(FILE_NAME)

   
    wb = load_workbook(FILE_NAME)
    sheet = wb.active

    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        recorded_name, recorded_date, _ = row

        if recorded_name == name and recorded_date == date:
            print(f"[!] {name} already marked today")
            return False

   
    sheet.append([name, date, time])
    wb.save(FILE_NAME)

    print(f"[✔] Attendance saved (Excel) for {name}")
    return True
